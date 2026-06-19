#!/usr/bin/env python3
"""
Build map data for advisory board members from the workbook.

Outputs:
- content/university-locations.csv (cached geocoded locations)
- content/advisory-map.json (map markers + year filter data)

Usage:
  python3 scripts/build-advisory-map.py
"""

from __future__ import annotations

import csv
import json
import ssl
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "List of Advisory Board Members 2021-2026.xlsx"
LOCATION_CACHE = ROOT / "content" / "university-locations.csv"
OUTPUT_JSON = ROOT / "content" / "advisory-map.json"

USER_AGENT = "WISE-Website-MapBuilder/1.0 (contact: wiseconnections)"
SSL_CONTEXT = ssl._create_unverified_context()


def normalize(value: object) -> str:
    return str(value or "").strip()


def load_rows() -> list[dict[str, str]]:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet_name = "List" if "List" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    headers = [normalize(c.value).lower() for c in ws[1]]

    def idx(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError as exc:
            raise ValueError(f"Missing column '{name}' in {sheet_name}") from exc

    first_i = idx("first name")
    last_i = idx("last name")
    year_i = idx("board year")
    university_i = idx("university represented")

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        university = normalize(row[university_i])
        year = normalize(row[year_i])
        if not university or not year:
            continue
        full_name = f"{normalize(row[first_i])} {normalize(row[last_i])}".strip()
        rows.append(
            {
                "name": " ".join(full_name.split()),
                "year": year,
                "university": " ".join(university.split()),
            }
        )
    return rows


def read_cache() -> dict[str, dict[str, str]]:
    if not LOCATION_CACHE.exists():
        return {}
    cache: dict[str, dict[str, str]] = {}
    with LOCATION_CACHE.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            university = normalize(row.get("university"))
            if not university:
                continue
            cache[university.casefold()] = {
                "university": university,
                "city": normalize(row.get("city")),
                "state": normalize(row.get("state")),
                "latitude": normalize(row.get("latitude")),
                "longitude": normalize(row.get("longitude")),
                "display_name": normalize(row.get("display_name")),
            }
    return cache


def geocode_university(name: str) -> dict[str, str] | None:
    query = urllib.parse.urlencode(
        {
            "q": f"{name}, USA",
            "format": "jsonv2",
            "limit": 1,
            "countrycodes": "us",
        }
    )
    url = f"https://nominatim.openstreetmap.org/search?{query}"
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=20, context=SSL_CONTEXT) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    if not payload:
        return None
    best = payload[0]

    address = best.get("address", {})
    city = (
        address.get("city")
        or address.get("town")
        or address.get("village")
        or address.get("municipality")
        or ""
    )
    state = address.get("state") or ""

    return {
        "university": name,
        "city": city,
        "state": state,
        "latitude": str(best.get("lat", "")),
        "longitude": str(best.get("lon", "")),
        "display_name": best.get("display_name", ""),
    }


def resolve_locations(universities: list[str], cache: dict[str, dict[str, str]]) -> dict[str, dict[str, str]]:
    resolved: dict[str, dict[str, str]] = {}
    misses = 0
    for university in sorted(universities, key=str.casefold):
        key = university.casefold()
        cached = cache.get(key)
        if cached and cached.get("latitude") and cached.get("longitude"):
            resolved[university] = cached
            continue

        try:
            result = geocode_university(university)
            if result:
                resolved[university] = result
                cache[key] = result
            else:
                misses += 1
                resolved[university] = {
                    "university": university,
                    "city": "",
                    "state": "",
                    "latitude": "",
                    "longitude": "",
                    "display_name": "",
                }
        except Exception:
            misses += 1
            resolved[university] = {
                "university": university,
                "city": "",
                "state": "",
                "latitude": "",
                "longitude": "",
                "display_name": "",
            }
        time.sleep(1.0)

    if misses:
        print(f"Could not geocode {misses} universities (left blank in cache).")
    return resolved


def write_cache(cache: dict[str, dict[str, str]]) -> None:
    LOCATION_CACHE.parent.mkdir(parents=True, exist_ok=True)
    rows = sorted(cache.values(), key=lambda r: r["university"].casefold())
    with LOCATION_CACHE.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["university", "city", "state", "latitude", "longitude", "display_name"],
        )
        writer.writeheader()
        writer.writerows(rows)


def build_output(rows: list[dict[str, str]], locations: dict[str, dict[str, str]]) -> dict[str, object]:
    years = sorted({r["year"] for r in rows}, key=int)
    by_year: dict[str, list[dict[str, object]]] = defaultdict(list)
    counts: dict[str, int] = defaultdict(int)

    for row in rows:
        loc = locations[row["university"]]
        lat = loc.get("latitude")
        lon = loc.get("longitude")
        if not lat or not lon:
            continue
        marker = {
            "name": row["name"],
            "university": row["university"],
            "year": row["year"],
            "city": loc.get("city", ""),
            "state": loc.get("state", ""),
            "latitude": float(lat),
            "longitude": float(lon),
        }
        by_year[row["year"]].append(marker)
        counts[row["year"]] += 1

    all_markers = [marker for year in years for marker in by_year.get(year, [])]
    return {
        "years": years,
        "countsByYear": counts,
        "allMarkers": all_markers,
        "markersByYear": by_year,
        "generatedAt": time.strftime("%Y-%m-%d %H:%M:%S"),
    }


def main() -> int:
    rows = load_rows()
    universities = sorted({r["university"] for r in rows}, key=str.casefold)
    cache = read_cache()
    locations = resolve_locations(universities, cache)
    write_cache(cache)

    output = build_output(rows, locations)
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(output, indent=2), encoding="utf-8")

    print(f"Rows read: {len(rows)}")
    print(f"Universities: {len(universities)}")
    print(f"Map markers: {len(output['allMarkers'])}")
    print(f"Wrote {LOCATION_CACHE.relative_to(ROOT)} and {OUTPUT_JSON.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
