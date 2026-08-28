#!/usr/bin/env python3
"""
Sync website content from Excel files in Resources/ into content/*.csv

Run after editing spreadsheets:
  python3 scripts/sync-from-excel.py

Requires: pip install openpyxl
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CONTENT = ROOT / "content"
RESOURCES = ROOT / "Resources"


def write_csv(
    path: Path,
    fieldnames: list[str],
    rows: list[dict],
    *,
    lineterminator: str = "\r\n",
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator=lineterminator)
        writer.writeheader()
        writer.writerows(rows)


def sync_chapters() -> None:
    path = RESOURCES / "Chapters.xlsx"
    if not path.exists():
        print("Skip chapters: Chapters.xlsx not found")
        return

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
        chapter_index = next(
            (index for index, header in enumerate(headers) if header.startswith("chapter")),
            0,
        )
        social_index = next(
            (index for index, header in enumerate(headers) if header == "social media"),
            None,
        )

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or chapter_index >= len(row) or not row[chapter_index]:
                continue

            university = str(row[chapter_index]).strip()
            if university.casefold() == "not in website yet":
                break

            # Keep public labels concise while preserving the workbook as the source.
            if "(Merged in Supply Chain Club" in university:
                university = university.split("(", 1)[0].strip()
            university = university.replace("University of Oklahama", "University of Oklahoma")

            social_url = ""
            if social_index is not None and social_index < len(row) and row[social_index]:
                candidate = str(row[social_index]).strip()
                if candidate.startswith(("https://", "http://")):
                    social_url = candidate

            rows.append({"university": university, "social_url": social_url})

        if rows:
            break

    if not rows:
        print("Skip chapters: no rows in Chapters.xlsx (existing chapters.csv kept)")
        return
    write_csv(
        CONTENT / "chapters.csv",
        ["university", "social_url"],
        rows,
        lineterminator="\n",
    )
    print(f"Wrote {len(rows)} chapters")


def sync_past_advisory() -> None:
    path = RESOURCES / "Advisory Board.xlsx"
    if not path.exists():
        print("Skip past advisory board: Advisory Board.xlsx not found")
        return

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    rows = []
    for year in wb.sheetnames:
        if not str(year).isdigit():
            continue
        ws = wb[year]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            photo = f"Resources/Photo keeping/AdvisoryBoard_{year}/{name}.jpg"
            rows.append({"year": str(year), "name": name, "photo": photo})

    write_csv(CONTENT / "past-advisory-board.csv", ["year", "name", "photo"], rows)
    print(f"Wrote {len(rows)} past advisory board entries")


def sync_conferences() -> None:
    path = RESOURCES / "Conferences.xlsx"
    if not path.exists():
        print("Skip conferences: Conferences.xlsx not found")
        return

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    rows = []
    for year in wb.sheetnames:
        if not str(year).isdigit():
            continue
        ws = wb[year]
        data = {}
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            key = str(row[0]).strip().lower()
            val = row[1] if len(row) > 1 else ""
            if key.startswith("date"):
                data["dates"] = str(val or "").strip()
            elif key.startswith("theme"):
                data["theme"] = str(val or "").strip()
            elif key.startswith("annual"):
                data["annual"] = str(val or "").strip()
            elif key.startswith("location"):
                data["location"] = str(val or "").strip()
            elif "university" in key:
                data["universities"] = str(val or "").strip()

        rows.append(
            {
                "year": str(year),
                "annual": data.get("annual", ""),
                "dates": data.get("dates", ""),
                "theme": data.get("theme", ""),
                "location": data.get("location", ""),
                "universities": data.get("universities", ""),
                "program_pdf": "",
                "sponsors": "",
            }
        )

    existing_pdfs = {}
    out_path = CONTENT / "conferences.csv"
    if out_path.exists():
        with out_path.open(encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("program_pdf"):
                    existing_pdfs[row["year"]] = row["program_pdf"]

    for row in rows:
        if not row.get("program_pdf") and row["year"] in existing_pdfs:
            row["program_pdf"] = existing_pdfs[row["year"]]

    rows.sort(key=lambda r: int(r["year"]), reverse=True)
    write_csv(
        CONTENT / "conferences.csv",
        ["year", "annual", "dates", "theme", "location", "universities", "program_pdf", "sponsors"],
        rows,
    )
    print(f"Wrote {len(rows)} conferences")


def main() -> int:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        return 1

    sync_chapters()
    sync_past_advisory()
    sync_conferences()
    print("Done. Refresh the site in your browser.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
