#!/usr/bin/env python3
"""
Fallback: refresh the University tab with unique names from the List sheet.

Use this only if your Excel version does not support dynamic array formulas
(UNIQUE / FILTER / SORT). For Microsoft 365, the University tab updates itself.

  python3 scripts/update-university-list.py
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "List of Advisory Board Members 2021-2026.xlsx"


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        return 1

    if not WORKBOOK.exists():
        print(f"Workbook not found: {WORKBOOK}", file=sys.stderr)
        return 1

    wb = load_workbook(WORKBOOK)
    source = "List" if "List" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[source]

    uni_col = None
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value is not None else ""
        if val.casefold() == "university represented":
            uni_col = cell.column
            break
    if uni_col is None:
        print("Column 'University Represented' not found.", file=sys.stderr)
        return 1

    seen: set[str] = set()
    unis: list[str] = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, uni_col).value
        if val is None:
            continue
        name = str(val).strip()
        if not name:
            continue
        key = name.casefold()
        if key not in seen:
            seen.add(key)
            unis.append(name)
    unis.sort(key=lambda s: s.casefold())

    uws = wb["University"] if "University" in wb.sheetnames else wb.create_sheet("University")
    for row in uws.iter_rows(min_row=1, max_row=max(uws.max_row, 1), min_col=1, max_col=1):
        for cell in row:
            cell.value = None

    uws["A1"] = "University"
    for i, name in enumerate(unis, start=2):
        uws.cell(i, 1, name)

    wb.save(WORKBOOK)
    print(f"Wrote {len(unis)} unique universities to University tab.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
